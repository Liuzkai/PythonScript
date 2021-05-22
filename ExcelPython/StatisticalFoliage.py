import openpyxl
import os
import csv
from openpyxl import Workbook
import math
import numpy as np
import pandas as pd


class MyCsvFile:

    def __init__(self, path):
        self.name = os.path.basename(path).split(".")[0]
            # path.split("\\")[-1].split(".")[0]
        self.index = int(self.name.split("_")[-2])
        self.path = path

    def __gt__(self, other):
        return self.index > other.index

    def __le__(self, other):
        return self.index < other.index

    def __eq__(self, other):
        return self.index == other.index

    def __str__(self):
        msg = "{} : {}".format(self.name, self.index)
        return msg


class StatisticalFoliage:

    def __init__(self, path):
        self.wb = Workbook()
        self.path = path
        self.csvfiles = self.find_sub_path(path)
        self.LOD0 = []
        self.LOD1 = []
        self.LOD2 = []
        self.LOD_0_Instance = []
        self.LOD_1_Instance = []
        self.LOD_2_Instance = []
        self.batch = []
        self.batch_instance = []
        self.combine_csv()
        self.statistical()
        self.save_xlsx()
        print("{} run end".format(type(self)))

    def find_sub_path(self, path):
        csv_list = []
        temp_files = os.listdir(path)
        for file in temp_files:
            full_path = os.path.join(path, file)
            if os.path.isfile(full_path) and os.path.splitext(full_path)[1] == '.csv':
                csv_list.append(MyCsvFile(full_path))
            elif os.path.isdir(full_path):
                csv_list.extend(self.find_sub_path(full_path))
        return csv_list

    def statistical(self):
        for ws in self.wb.worksheets[1:]:
            LOD0 = 0
            LOD1 = 0
            LOD2 = 0
            LOD_0_Instance = 0
            LOD_1_Instance = 0
            LOD_2_Instance = 0
            batch = 0
            batch_instance = 0
            for cell in ws['A']:
                if str(cell.value) == "STATGROUP_FoliageDetailTriangle":
                    row = cell.row
                    if "Batch" in str(ws.cell(row=row, column=2).value):
                        batch += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD0'):
                        LOD0 += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD1'):
                        LOD1 += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD2'):
                        LOD2 += int(ws.cell(row=row, column=3).value)
                elif str(cell.value) == "STATGROUP_FoliageDetailInstance":
                    row = cell.row
                    if "Batch" in str(ws.cell(row=row, column=2).value):
                        batch_instance += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD0'):
                        LOD_0_Instance += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD1'):
                        LOD_1_Instance += int(ws.cell(row=row, column=3).value)
                    elif str(ws.cell(row=row, column=2).value).endswith('LOD2'):
                        LOD_2_Instance += int(ws.cell(row=row, column=3).value)
            self.LOD0.append(LOD0)
            self.LOD1.append(LOD1)
            self.LOD2.append(LOD2)
            self.LOD_0_Instance.append(LOD_0_Instance)
            self.LOD_1_Instance.append(LOD_1_Instance)
            self.LOD_2_Instance.append(LOD_2_Instance)
            self.batch.append(batch)
            self.batch_instance.append(batch_instance)

    def combine_csv(self):
        # combine csv to xlsx
        csvfilelist = sorted(self.csvfiles)
        for csvfile in csvfilelist:
            with open(csvfile.path, "r", encoding='utf-8') as f:
                content = csv.reader(f, delimiter=',')
                name = str(csvfile.index)
                ws = self.wb.create_sheet(name)
                for row in content:
                    ws.append(row)
                # convert string number to number
                for i in range(2, ws.max_row + 1):
                    if ws.cell(row=i, column=3).value.isnumeric():
                        ws.cell(row=i, column=3, value=int(ws.cell(row=i, column=3).value))

    def combine_csv_with_sort(self):
        csvfilelist = sorted(self.csvfiles)
        for csvfile in csvfilelist:
            data = np.genfromtxt(csvfile.path, delimiter=",")



    def save_xlsx(self):
        ws = self.wb.worksheets[0]
        ws.cell(row=1, column=1, value="Camera")
        ws.cell(row=1, column=2, value="LOD0")
        ws.cell(row=1, column=3, value="LOD1")
        ws.cell(row=1, column=4, value="LOD2")
        ws.cell(row=1, column=5, value="Batch")
        ws.cell(row=1, column=6, value="LOD0 Instance")
        ws.cell(row=1, column=7, value="LOD1 Instance")
        ws.cell(row=1, column=8, value="LOD2 Instance")
        ws.cell(row=1, column=9, value="Batch Instance")
        ws.cell(row=1, column=10, value="Triangle Total")
        ws.cell(row=1, column=11, value="Instance Total")
        # the total data row
        data_row = len(self.LOD0)
        # set Camera Name
        for i in range(data_row):
            ws.cell(row=i + 2, column=1, value="camera {}".format(i + 1))
        # set Statistical Data
        data = (self.LOD0 + self.LOD_0_Instance
                + self.LOD1 + self.LOD_1_Instance
                + self.LOD2 + self.LOD_2_Instance
                + self.batch + self.batch_instance)
        for i, val in enumerate(data):
            ws.cell(row=(i % data_row + 2), column=math.floor(i / data_row + 2), value=val)
        # set Total Data
        for i in range(data_row):
            ws["J{}".format(i + 2)] = '=SUM(B{}:E{})'.format(i + 2, i + 2)
            ws["K{}".format(i + 2)] = '=SUM(F{}:I{})'.format(i + 2, i + 2)
        # save file
        name = self.path + "/result.xlsx"
        self.wb.save(name)


StatisticalFoliage(r"/Users/liuzhongkai/Documents/Git/PythonScript/ExcelPython")
# data = pd.read_csv(r"D:\NExTWorkSpace\Arkm_ArtBranch\ArkGame\Saved\Screenshots\Windows\seq04\csv\viewport_seq04_1_1.csv")
# b = data.sort_values(by=['StatGroupName'])
# print(b)