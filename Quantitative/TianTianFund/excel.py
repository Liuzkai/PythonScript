import requests    #requests是HTTP库
import re
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
from bs4 import BeautifulSoup as bs   #bs:通过解析文档为用户提供需要抓取的数据
import os
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf8') #改变标准输出的默认编码

#我们开始利用requests.get（）来获取网页并利用bs4解析网页：
def getData(src):

    html = requests.get(src).content    # requests.get(src)返回的是状态码<Response [200]>，加上.content以字节形式（二进制返回数据。   和前端一样，分为get post等  http://www.cnblogs.com/ranxf/p/7808537.html
    soup = bs(html,'html.parser')   # lxml改为html.parser, 解析器解析字节形式的数据，得到完整的类似页面的html代码结构的数据
    print(soup)

    global ws
    Name = []
    Introductions = []
    introductions = soup.find_all("a",class_="book-item-name")
    nameList = soup.find_all("a",class_="author")
    print (nameList)
    for name in nameList:
        print (name.text)
        Name.append(name.text)
    for introduction in introductions:
        Introductions.append(introduction.text)
    for i in range(len(Name)):
        ws.append([Name[i],Introductions[i]])

if __name__ == '__main__':
    #   读取存在的Excel表测试
    #     wb = load_workbook('t est.xlsx') #加载存在的Excel表
    #     a_sheet = wb.get_sheet_by_name('Sheet1') #根据表名获取表对象
    #     for row in a_sheet.rows: #遍历输出行数据
    #         for cell in row: #每行的 每一个单元格
    #             print cell.value,

    #  创建Excel表并写入数据
    wb = workbook.Workbook()  # 创建Excel对象
    ws = wb.active  # 获取当前正在操作的表对象
    # 往表中写入标题行,以列表形式写入！
    ws.append(['角色名字', '票数'])
    src = 'http://www.lrts.me/book/category/3058'
    getData(src)
    wb.save('qinshi.xlsx')  # 存入所有信息后，保存为filename.xlsx